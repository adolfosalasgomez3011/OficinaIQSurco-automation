from __future__ import annotations

import csv
import re
import shutil
import tempfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, List
from urllib.parse import quote

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


BASE_DIR = Path(__file__).resolve().parent.parent
SOURCE_WORKBOOK = BASE_DIR / "FacebookAdsCampaing" / "CRMIQSurcov1.xlsx"
TRACKER_CSV = BASE_DIR / "LeadAutomation" / "lead_followup_tracker.csv"
QUEUE_WORKBOOK = BASE_DIR / "LeadAutomation" / "WhatsAppFollowupQueue.xlsx"

OWNER = "Adolfo Salas"
DEFAULT_STAGE = "New inbound"
DEFAULT_INTENT = "no claro"
DEFAULT_NEXT_ACTION = "Enviar primer WhatsApp"
DEFAULT_PRIORITY = "media"
CONTACTED_STATUSES = {"sent", "replied", "qualified", "scheduled", "closed"}


@dataclass
class Lead:
    lead_id: str
    created_time: str
    campaign_name: str
    platform: str
    email: str
    full_name: str
    phone_number: str
    lead_status: str


def normalize_phone(raw_phone: str) -> str:
    cleaned = (raw_phone or "").replace("p:", "").strip()
    digits = re.sub(r"\D", "", cleaned)
    if not digits:
        return ""
    if digits.startswith("00"):
        digits = digits[2:]
    return digits


def first_name(full_name: str) -> str:
    base = (full_name or "").strip().split()
    return base[0].title() if base else ""


def is_test_lead(full_name: str, phone_number: str) -> bool:
    sample = f"{full_name} {phone_number}".lower()
    return "<test lead:" in sample or "dummy data" in sample


def build_initial_message(name: str) -> str:
    greeting_name = first_name(name)
    greeting = f"Hola {greeting_name}," if greeting_name else "Hola,"
    return (
        f"{greeting} gracias por tu interés en las oficinas de IQ Surco. "
        "Tengo disponibles opciones en Av. La Encalada y te puedo compartir video, "
        "precios y horarios de visita. Me confirmas si estás buscando alquiler, compra o ambas opciones?"
    )


def load_meta_leads(workbook_path: Path) -> List[Lead]:
    try:
        workbook = load_workbook(workbook_path, data_only=True)
    except PermissionError:
        # Excel can lock the workbook; read from a temporary copy instead.
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
            temp_path = Path(temp_file.name)
        try:
            shutil.copy2(workbook_path, temp_path)
            workbook = load_workbook(temp_path, data_only=True)
        finally:
            try:
                temp_path.unlink(missing_ok=True)
            except OSError:
                pass

    sheet = workbook["Clients"]
    headers = [cell.value for cell in sheet[1]]
    header_index = {header: index for index, header in enumerate(headers)}

    leads: List[Lead] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        lead = Lead(
            lead_id=str(row[header_index["id"]] or "").strip(),
            created_time=str(row[header_index["created_time"]] or "").strip(),
            campaign_name=str(row[header_index["campaign_name"]] or "").strip(),
            platform=str(row[header_index["platform"]] or "").strip(),
            email=str(row[header_index["email"]] or "").strip(),
            full_name=str(row[header_index["full_name"]] or "").strip(),
            phone_number=str(row[header_index["phone_number"]] or "").strip(),
            lead_status=str(row[header_index["lead_status"]] or "").strip(),
        )
        if not lead.lead_id or not lead.full_name:
            continue
        if is_test_lead(lead.full_name, lead.phone_number):
            continue
        leads.append(lead)

    return leads


def load_tracker(tracker_path: Path) -> Dict[str, Dict[str, str]]:
    if not tracker_path.exists():
        return {}

    with tracker_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return {row["lead_id"]: row for row in reader if row.get("lead_id")}


def write_tracker(rows: List[Dict[str, str]], tracker_path: Path) -> None:
    fieldnames = [
        "lead_id",
        "created_time",
        "full_name",
        "phone_number",
        "email",
        "platform",
        "campaign_name",
        "stage",
        "intent",
        "outbound_status",
        "last_message_draft",
        "last_contact_at",
        "next_action",
        "followup_at",
        "owner",
        "priority",
        "notes",
    ]
    with tracker_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def merge_rows(leads: List[Lead], existing_tracker: Dict[str, Dict[str, str]]) -> List[Dict[str, str]]:
    phone_contacted_index: Dict[str, Dict[str, str]] = {}
    for tracked in existing_tracker.values():
        status = (tracked.get("outbound_status") or "").strip().lower()
        phone = normalize_phone(tracked.get("phone_number") or "")
        if status not in CONTACTED_STATUSES or not phone:
            continue
        previous = phone_contacted_index.get(phone)
        if not previous:
            phone_contacted_index[phone] = tracked
            continue

        prev_date = previous.get("last_contact_at") or previous.get("created_time") or ""
        curr_date = tracked.get("last_contact_at") or tracked.get("created_time") or ""
        if curr_date > prev_date:
            phone_contacted_index[phone] = tracked

    merged_rows: List[Dict[str, str]] = []
    today_iso = datetime.now().strftime("%Y-%m-%d")
    for lead in sorted(leads, key=lambda item: item.created_time):
        normalized = normalize_phone(lead.phone_number)
        existing = existing_tracker.get(lead.lead_id, {})
        draft_message = existing.get("last_message_draft") or build_initial_message(lead.full_name)

        auto_status = existing.get("outbound_status") or "pending"
        auto_next_action = existing.get("next_action") or DEFAULT_NEXT_ACTION
        auto_notes = existing.get("notes") or ""

        if not existing and normalized in phone_contacted_index:
            previous = phone_contacted_index[normalized]
            previous_lead = previous.get("lead_id") or "sin lead_id"
            previous_status = previous.get("outbound_status") or "sin status"
            previous_contact = previous.get("last_contact_at") or previous.get("created_time") or "sin fecha"
            auto_status = "already_contacted"
            auto_next_action = "Revisar duplicado por telefono"
            auto_notes = (
                "Auto-flag: telefono ya contactado. "
                f"lead_id previo={previous_lead}, status={previous_status}, fecha={previous_contact}."
            )

        last_contact_at = existing.get("last_contact_at") or ""
        if auto_status.lower() in CONTACTED_STATUSES and not last_contact_at:
            # Auto-stamp first contact date when status indicates the lead was contacted.
            last_contact_at = today_iso

        merged_rows.append(
            {
                "lead_id": lead.lead_id,
                "created_time": lead.created_time,
                "full_name": lead.full_name,
                "phone_number": normalized,
                "email": lead.email,
                "platform": lead.platform or ("organic" if lead.lead_status == "TRUE" else "meta"),
                "campaign_name": lead.campaign_name,
                "stage": existing.get("stage") or DEFAULT_STAGE,
                "intent": existing.get("intent") or DEFAULT_INTENT,
                "outbound_status": auto_status,
                "last_message_draft": draft_message,
                "last_contact_at": last_contact_at,
                "next_action": auto_next_action,
                "followup_at": existing.get("followup_at") or "",
                "owner": existing.get("owner") or OWNER,
                "priority": existing.get("priority") or DEFAULT_PRIORITY,
                "notes": auto_notes,
            }
        )
    return merged_rows


def build_whatsapp_url(phone_number: str, message: str) -> str:
    return f"https://wa.me/{phone_number}?text={quote(message)}"


def write_queue_workbook(rows: List[Dict[str, str]], workbook_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Pending"

    headers = [
        "lead_id",
        "created_time",
        "full_name",
        "platform",
        "campaign_name",
        "phone_number",
        "intent",
        "stage",
        "outbound_status",
        "next_action",
        "priority",
        "action_click_send",
        "message_ready_to_send",
        "last_contact_at",
        "followup_at",
        "notes",
    ]
    sheet.append(headers)

    header_font = Font(bold=True)
    for cell in sheet[1]:
        cell.font = header_font

    row_index = 2
    for row in rows:
        if row["outbound_status"].lower() not in {"pending", "drafted"}:
            continue

        message = row["last_message_draft"]
        sheet.cell(row=row_index, column=1, value=row["lead_id"])
        sheet.cell(row=row_index, column=2, value=row["created_time"])
        sheet.cell(row=row_index, column=3, value=row["full_name"])
        sheet.cell(row=row_index, column=4, value=row["platform"])
        sheet.cell(row=row_index, column=5, value=row["campaign_name"])
        sheet.cell(row=row_index, column=6, value=row["phone_number"])
        sheet.cell(row=row_index, column=7, value=row["intent"])
        sheet.cell(row=row_index, column=8, value=row["stage"])
        sheet.cell(row=row_index, column=9, value=row["outbound_status"])
        sheet.cell(row=row_index, column=10, value=row["next_action"])
        sheet.cell(row=row_index, column=11, value=row["priority"])
        link_cell = sheet.cell(row=row_index, column=12, value="Open Chat + Send")
        link_cell.hyperlink = build_whatsapp_url(row["phone_number"], message)
        link_cell.style = "Hyperlink"
        sheet.cell(row=row_index, column=13, value=message)
        sheet.cell(row=row_index, column=14, value=row["last_contact_at"])
        sheet.cell(row=row_index, column=15, value=row["followup_at"])
        sheet.cell(row=row_index, column=16, value=row["notes"])
        row_index += 1

    widths = {
        "A": 22,
        "B": 25,
        "C": 28,
        "D": 10,
        "E": 30,
        "F": 18,
        "G": 12,
        "H": 18,
        "I": 16,
        "J": 24,
        "K": 12,
        "L": 20,
        "M": 85,
        "N": 22,
        "O": 22,
        "P": 30,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    workbook.save(workbook_path)


def main() -> None:
    leads = load_meta_leads(SOURCE_WORKBOOK)
    tracker = load_tracker(TRACKER_CSV)
    merged_rows = merge_rows(leads, tracker)
    write_tracker(merged_rows, TRACKER_CSV)

    queue_output_path = QUEUE_WORKBOOK
    try:
        write_queue_workbook(merged_rows, queue_output_path)
    except PermissionError:
        # If Excel has the workbook open, save a fresh file with a timestamp suffix.
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        queue_output_path = QUEUE_WORKBOOK.with_name(f"WhatsAppFollowupQueue_{timestamp}.xlsx")
        write_queue_workbook(merged_rows, queue_output_path)

    pending_count = sum(1 for row in merged_rows if row["outbound_status"].lower() in {"pending", "drafted"})
    already_contacted_count = sum(1 for row in merged_rows if row["outbound_status"].lower() == "already_contacted")
    print(
        f"Synced {len(merged_rows)} leads. Pending WhatsApp follow-ups: {pending_count}. "
        f"Already contacted duplicates: {already_contacted_count}. "
        f"Tracker: {TRACKER_CSV.name}. Queue: {queue_output_path.name}."
    )


if __name__ == "__main__":
    main()